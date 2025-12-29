"""
Excel Generator Module
Handles Excel file generation with email data
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict
import io


class ExcelGenerator:
    """Handles Excel file generation with formatting"""
    
    def __init__(self):
        """Initialize the Excel workbook"""
        self.workbook = None
        self.worksheet = None
    
    def create_workbook(self):
        """Create a new Excel workbook"""
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Emails"
    
    def add_headers(self, headers: List[str]):
        """
        Add header row to the worksheet
        
        Args:
            headers (List[str]): List of header titles
        """
        if not self.worksheet:
            self.create_workbook()
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = header
            
            # Style headers
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Freeze the header row
        self.worksheet.freeze_panes = "A2"
    
    def add_emails(self, emails: List[str]):
        """
        Add emails to the worksheet with header row
        
        Args:
            emails (List[str]): List of email addresses
        """
        if not self.worksheet:
            self.create_workbook()
        
        # Always add header in the first row
        self.add_headers(["Email"])
        
        # Add emails starting from row 2 (after header)
        for row_num, email in enumerate(emails, 2):
            cell = self.worksheet.cell(row=row_num, column=1)
            cell.value = email
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    def add_emails_with_metadata(self, data: List[Dict]):
        """
        Add emails with additional metadata (name, company, etc.)
        
        Args:
            data (List[Dict]): List of dictionaries with email and metadata
                               Example: [{'email': 'john@example.com', 'name': 'John', 'company': 'ABC Inc'}]
        """
        if not data:
            return
        
        if not self.worksheet:
            self.create_workbook()
        
        # Extract headers from first record
        headers = list(data[0].keys())
        self.add_headers(headers)
        
        # Add data rows
        for row_num, record in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=row_num, column=col_num)
                cell.value = record.get(header, "")
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    def auto_fit_columns(self):
        """Auto-fit column widths based on content"""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def save_to_bytes(self) -> bytes:
        """
        Save the workbook to bytes (for Streamlit download)
        
        Returns:
            bytes: Excel file content as bytes
        """
        if not self.workbook:
            raise ValueError("No workbook created. Add data first.")
        
        # Auto-fit columns before saving
        self.auto_fit_columns()
        
        # Save to bytes buffer
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.getvalue()
    
    def save_to_file(self, filepath: str):
        """
        Save the workbook to a file
        
        Args:
            filepath (str): Path to save the Excel file
        """
        if not self.workbook:
            raise ValueError("No workbook created. Add data first.")
        
        # Auto-fit columns before saving
        self.auto_fit_columns()
        
        self.workbook.save(filepath)
        print(f"Excel file saved to: {filepath}")


def generate_excel_from_emails(emails: List[str]) -> bytes:
    """
    Convenience function to generate Excel file from emails
    
    Args:
        emails (List[str]): List of email addresses
        
    Returns:
        bytes: Excel file content as bytes
    """
    generator = ExcelGenerator()
    generator.create_workbook()
    generator.add_emails(emails)
    return generator.save_to_bytes()


def generate_excel_from_data(data: List[Dict]) -> bytes:
    """
    Convenience function to generate Excel file with metadata
    
    Args:
        data (List[Dict]): List of dictionaries with email and other fields
        
    Returns:
        bytes: Excel file content as bytes
    """
    generator = ExcelGenerator()
    generator.create_workbook()
    generator.add_emails_with_metadata(data)
    return generator.save_to_bytes()
